"""
============================================================
Script:   product_image_auditor.py
Author:   Stanley Carter
Purpose:  Audits product image availability on the Amplience
          CDN for each SKU in the master sample list workbook.
          Writes SHOT or NEEDS PHOTO status back to the
          spreadsheet and links confirmed images for review.

HOW IT WORKS:
  1. Loads the master Excel workbook
  2. For each target department tab, reads style and color
     numbers from every data row
  3. Constructs a CDN URL per SKU:
       https://i8.amplience.net/s/scvl/{style}_{color}_SET
  4. Makes concurrent HTTP requests (20 workers) to check
     whether a real image exists at each URL
  5. Distinguishes real images from placeholder images by
     checking response payload size against a known
     placeholder threshold
  6. Writes results back to the workbook:
       - Column T: "SHOT" or "NEEDS PHOTO"
       - Column U: clickable HYPERLINK formula if image exists
  7. Saves after each tab so a mid-run crash loses no work

PERFORMANCE:
  ThreadPoolExecutor with 20 workers allows concurrent
  checking across all rows in a tab. Progress is logged
  every 100 rows. Runtime scales with tab size and network.

DEPENDENCIES:
  pip install openpyxl requests
============================================================
"""

import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

# -------------------------------------------------------
# CONFIGURATION — update paths and column numbers as needed
# -------------------------------------------------------

FILE_PATH = r"C:\Users\scarter\OneDrive - Shoe Carnival, Inc\Sample Request Lists\Master Sample List 2026.xlsx"

BASE_URL = "https://i8.amplience.net/s/scvl/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# Tabs to audit — must match exact sheet names in the workbook
# Note: "BGRP Summary" excluded — different structure, insufficient columns
TARGET_TABS = [
    "M ATH",
    "W ATH",
    "C ATH",
    "MENS",
    "WOMENS",
    "CHILDRENS",
    "ACCESSORIES",
]

# Column positions (1-indexed, matching workbook layout)
COL_STYLE  = 12   # Column L: Ecom Style Number
COL_COLOR  = 13   # Column M: Ecom Color Number
COL_STATUS = 20   # Column T: Image status output (SHOT / NEEDS PHOTO)
COL_LINK   = 21   # Column U: Hyperlink to confirmed image

# Known byte size of the Amplience placeholder image.
# Responses at or below this size are treated as "no real image."
PLACEHOLDER_SIZE = 12302


# -------------------------------------------------------
# IMAGE CHECK — returns True if a real image exists at url
# -------------------------------------------------------

def check_image(url):
    """
    Makes a streaming GET request to the CDN URL.
    Reads up to PLACEHOLDER_SIZE + 100 bytes of the response.
    If payload exceeds the placeholder threshold, a real image exists.
    Returns False on any error or non-200 status.
    """
    try:
        r = requests.get(url, headers=HEADERS, timeout=5, stream=True)
        if r.status_code != 200:
            r.close()
            return False

        chunk = b""
        for data in r.iter_content(chunk_size=1024):
            chunk += data
            if len(chunk) > PLACEHOLDER_SIZE + 100:
                r.close()
                return True  # Real image — larger than placeholder

        r.close()
        return False  # Response too small — placeholder or empty

    except Exception:
        return False  # Network error, timeout, etc.


# -------------------------------------------------------
# MAIN — load workbook and audit each target tab
# -------------------------------------------------------

print("Loading workbook...")
wb = openpyxl.load_workbook(FILE_PATH)

for sheet_name in TARGET_TABS:

    # Skip if tab doesn't exist in this workbook version
    if sheet_name not in wb.sheetnames:
        print(f"\nSkipping '{sheet_name}' — tab not found.")
        continue

    print(f"\nAuditing tab: {sheet_name}")
    ws = wb[sheet_name]

    # Build list of (row_index, url) pairs for all valid data rows
    work = []
    for row in ws.iter_rows(min_row=3):  # Row 1-2 are headers
        row_idx = row[0].row

        # Skip rows without enough columns to read style/color
        if len(row) < COL_COLOR:
            continue

        style_val = row[COL_STYLE - 1].value
        color_val = row[COL_COLOR - 1].value

        # Skip empty rows
        if style_val is None or color_val is None:
            continue

        # Normalize: strip decimals (e.g. 12345.0 → "12345")
        style_num = str(style_val).split(".")[0].strip()
        color_num = str(color_val).split(".")[0].strip()

        # Skip blanks or corrupted values
        if not style_num or not color_num or "nan" in (style_num, color_num):
            continue

        url = f"{BASE_URL}{style_num}_{color_num}_SET"
        work.append((row_idx, url))

    if not work:
        print(f"  No valid data rows found — skipping.")
        continue

    print(f"  Checking {len(work)} rows...")

    # Concurrent image checks using thread pool
    results = {}
    lock = Lock()

    with ThreadPoolExecutor(max_workers=20) as pool:
        future_map = {
            pool.submit(check_image, url): (row_idx, url)
            for row_idx, url in work
        }
        done = 0
        for future in as_completed(future_map):
            row_idx, url = future_map[future]
            exists = future.result()
            with lock:
                results[row_idx] = (exists, url)
                done += 1
                if done % 100 == 0:
                    print(f"    {done}/{len(work)} checked...")

    # Write results back to workbook
    shot = needs = 0
    for row_idx, (exists, url) in results.items():
        if exists:
            ws.cell(row=row_idx, column=COL_STATUS).value = "SHOT"
            ws.cell(row=row_idx, column=COL_LINK).value   = f'=HYPERLINK("{url}","View Photo")'
            shot += 1
        else:
            ws.cell(row=row_idx, column=COL_STATUS).value = "NEEDS PHOTO"
            ws.cell(row=row_idx, column=COL_LINK).value   = ""
            needs += 1

    print(f"  SHOT: {shot} | NEEDS PHOTO: {needs}")

    # Save after every tab — protects completed work if a later tab crashes
    print(f"  Saving after {sheet_name}...")
    wb.save(FILE_PATH)
    print(f"  Saved.")

print("\nAll tabs complete.")
